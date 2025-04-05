from openpyxl import Workbook, load_workbook

libro = Workbook()

hoja = libro.active

hoja.title = "Datos"

libro.create_sheet("Ventas")

libro.active = 1

hoja2 = libro.active


datos = [
    ["Producto", "Cantidad", "Precio"],
    ["Manzanas", 50, 0.5],
    ["Naranjas", 30, 0.75]
]

filas= list(range(1,4))
columnas = ['A','B','C']

for f in filas:
    for c in columnas:
        celda = c+str(f)
        hoja2[celda]=datos[filas.index(f)][columnas.index(c)]

libro.save("mi_libro.xlsx")


libro2 = Workbook()

libro2 = load_workbook("mi_libro.xlsx")

libro2.active = 1

hoja2 = libro2.active

problema3 = list(hoja2)

for fila in problema3:
    for columna in fila:
        print(columna.value)











